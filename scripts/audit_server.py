#!/usr/bin/env python3.11
"""
audit_server.py — 慈恩資料遷移稽核 Web 工具

啟動:
    python3.11 scripts/audit_server.py
    然後開啟瀏覽器 http://localhost:8787
"""

import io
import json
import re
import sys
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("請先安裝: python3.11 -m pip install openpyxl fastapi uvicorn python-multipart", file=sys.stderr)
    sys.exit(1)

try:
    import uvicorn
    from fastapi import FastAPI, File, UploadFile
    from fastapi.responses import HTMLResponse, JSONResponse
except ImportError:
    print("請先安裝: python3.11 -m pip install fastapi uvicorn python-multipart", file=sys.stderr)
    sys.exit(1)

# ─── 靜態映射表 ───────────────────────────────────────────────────────────────

THERAPIST_MAP: dict[str, str] = {
    "呂孟育": "T001", "林紀宇": "T002", "林容蒂": "T003",
    "邱似齡": "T004", "蔡孟潔": "T005", "陳慧苓": "T006",
    "游子瑩": "T007", "葉邦彥": "T008", "鄭幼毅": "T009",
    "楊顯欽": "T010", "劉柏宏": "T011", "邱惟雅": "T012",
    "潘柔靜": "T013", "邱意祺": "T014", "劉彥君": "T015",
    "羅紀萱": "T016", "黃慧婷": "T017",
    "劉子瑩": "T018", "楊子賢": "T019", "林郁珊": "T020", "蕭雅文": "T021",
}

ROOM_CODES: set[str] = {
    "1A", "1B", "1C", "1D",
    "2A", "2B", "2C", "2D", "2E",
    "3A", "3B", "3C", "3D", "3E", "3F",
}

SELF_PAY_SOURCES: set[str] = {
    "自費", "自費伴侶", "自費(個別轉伴侶)", "自費(家族)", "家事自費",
    "教支轉自費", "澎湖教支轉自費", "",
}

INSTITUTION_MAP: dict[str, str] = {
    "教支": "教育支持方案", "台南市教支": "教育支持方案",
    "台積電": "台積電EAP", "台積EAP": "台積電EAP",
    "台積EAP伴侶": "台積電EAP", "蛹之生台積電": "台積電EAP",
    "台電": "台電", "奇美": "奇美", "衛生局": "衛生局",
    "台南市家暴性侵防治中心": "家暴性侵防治中心",
    "台南市政府人事處": "人事處", "人事處": "人事處",
    "南家扶": "家扶基金會", "家扶FAP": "家扶基金會",
    "善牧": "善牧基金會", "國軍": "國軍",
    "地方法院": "法院", "法院": "法院", "法院EAP": "法院",
    "女權家事": "女權", "家事(女權)": "女權", "更生": "更生",
    "職災中心": "職災", "警局": "警局", "台南市警局": "警局",
    "青壯": "青壯方案", "脆家": "脆弱家庭",
    "醫事": "醫事人員方案", "醫事人員方案": "醫事人員方案",
    "家防": "家庭防暴", "家防伴侶": "家庭防暴",
    "永仁": "永仁", "永觀": "永觀", "永觀長照": "永觀",
    "戀戀交友": "戀戀交友", "荷光轉案伴侶": "荷光",
    "伴侶招募": "?待確認", "伴家加久方案": "?待確認",
    "實習生方案": "?待確認", "15-30方案": "?待確認", "15-45": "?待確認",
}

TW_ID_RE = re.compile(r'^[A-Z][12]\d{8}$')
DATE_RE = re.compile(r'^(\d{2,4})[/.](\d{1,2})[/.](\d{1,2})$')
DATE_STRIP_RE = re.compile(r'[（(][^）)]+[）)]')
TIME_RE = re.compile(r'^\d{1,2}:\d{2}$')
_FULLWIDTH = str.maketrans("０１２３４５６７８９", "0123456789")

def parse_date(raw, field_hint: str = "") -> tuple[date | None, str | None]:
    if raw is None or raw == "":
        return None, None
    if isinstance(raw, (datetime, date)):
        return raw.date() if isinstance(raw, datetime) else raw, None
    s = str(raw).strip()
    m = DATE_RE.match(s)
    if not m:
        return None, f"無法解析日期格式：{s!r}"
    y, mo, d_ = int(m.group(1)), int(m.group(2)), int(m.group(3))
    if y < 200:
        y += 1911
    try:
        return date(y, mo, d_), None
    except ValueError as e:
        return None, f"日期值無效 {y}-{mo:02d}-{d_:02d}：{e}"

def parse_gender(raw) -> str | None:
    if raw is None or raw == "":
        return None
    s = str(raw).strip().translate(_FULLWIDTH)
    if s in ("1", "男"):
        return "male"
    if s in ("2", "女"):
        return "female"
    return "other"

def is_couple_name(name: str) -> bool:
    """偵測名字是否為伴侶案：含（伴侶）、(伴侶)，或兩人名以 、 分隔"""
    if not name:
        return False
    if "伴侶" in name:
        return True
    if "、" in name:
        return True
    return False

def split_couple_names(name: str) -> list[str]:
    """從伴侶名稱中提取個人名字"""
    clean = name.replace("(伴侶)", "").replace("（伴侶）", "").strip()
    if "、" in clean:
        return [n.strip() for n in clean.split("、") if n.strip()]
    return [clean] if clean else []

# ─── 讀取 xlsx ────────────────────────────────────────────────────────────────

def load_workbook_from_bytes(data: bytes) -> openpyxl.Workbook:
    return openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)

def read_sheet3(wb) -> list[dict]:
    """進行中個案完整資料"""
    ws = wb["進行中個案完整資料"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(c).strip() if c else "" for c in rows[0]]

    def col(row, name, default=None):
        try:
            idx = header.index(name)
            return row[idx] if idx < len(row) else default
        except ValueError:
            return default

    cases = []
    for i, row in enumerate(rows[1:], start=2):
        if not any(row):
            continue
        cases.append({
            "row": i,
            "case_code": str(col(row, "個案編號") or "").strip(),
            "name": str(col(row, "個案姓名") or "").strip(),
            "therapist": str(col(row, "接案心理師") or "").strip(),
            "funding_raw": str(col(row, "經費來源") or "").strip(),
            "gender_raw": col(row, "性別"),
            "national_id": str(col(row, "身分證字號") or "").strip(),
            "birth_raw": col(row, "出生日期"),
            "visit_raw": col(row, "進案日"),
            "phone": str(col(row, "連絡電話(手機)") or "").strip(),
            "phone_home": str(col(row, "連絡電話(市話)") or "").strip(),
            "emergency_contact": str(col(row, "緊急聯絡人") or "").strip(),
            "address": str(col(row, "地址") or "").strip(),
            "intake_year": str(col(row, "個案進案年份") or "").strip(),
            "appt_count": col(row, "2026預約次數") or 0,
        })
    return cases

def read_sheet5(wb) -> tuple[list[dict], int]:
    """預約明細_標準化名稱；returns (appts, skipped_count)"""
    ws = wb["預約明細_標準化名稱"]
    rows = list(ws.iter_rows(values_only=True))
    appts, skipped = [], 0
    for i, row in enumerate(rows[1:], start=2):
        if not any(row):
            continue
        case_code = str(row[7] or "").strip() if len(row) > 7 else ""
        raw_case_name = str(row[5] or "").strip()
        # Skip admin entries
        if not case_code or case_code in ("R",):
            skipped += 1
            continue
        appts.append({
            "row": i,
            "date_raw": str(row[0] or "").strip(),
            "start": str(row[1] or "").strip(),
            "end": str(row[2] or "").strip(),
            "room": str(row[3] or "").strip(),
            "therapist": str(row[4] or "").strip(),
            "case_name_raw": raw_case_name,
            "std_name": str(row[6] or "").strip() if len(row) > 6 else "",
            "case_code": case_code,
            "notes": str(row[8] or "").strip() if len(row) > 8 else "",
        })
    return appts, skipped

def read_problem_cases(wb) -> tuple[list[dict], list[dict], list[dict]]:
    ws = wb["需處理個案清單"]
    rows = list(ws.iter_rows(values_only=True))
    cat_c, cat_d, cat_b = [], [], []

    # Section 標題可能被 openpyxl 存檔時抹除，改用 header 出現次序判斷：
    # 第1個「個案姓名」header → Cat C；第2個 → Cat D；「跨年延續」後 or 第3個 → Cat B
    current, has_header = None, False
    header_count = 0

    for row in rows:
        if not any(row):
            continue
        cell1 = str(row[1] or "").strip()
        cell4 = str(row[4] or "").strip() if len(row) > 4 else ""

        # section 標題（未被抹除時）
        if "跨年延續" in cell1:
            current = "B"
            has_header = False
            continue
        if ("完全未列入" in cell1 and "≥3" in cell1):
            current = "C"
            has_header = False
            continue
        if ("完全未列入" in cell1 and "<3" in cell1):
            current = "D"
            has_header = False
            continue

        # header 列
        if cell1 == "個案姓名":
            header_count += 1
            has_header = True
            if "已建立年份" in cell4:   # Cat B 有獨特欄位
                current = "B"
            elif current is None or header_count == 1:
                current = "C"
            elif header_count == 2:
                current = "D"
            elif header_count >= 3:
                current = "B"
            continue

        # 資料列
        if current and has_header and row[0]:
            code = str(row[0]).strip()
            entry = {
                "code": code,
                "name": str(row[1] or "").strip(),
                "appt_count": row[2] or 0,
                "therapist": str(row[3] or "").strip(),
                "note": str(row[4] or "").strip() if len(row) > 4 else "",
                "category": current,
            }
            if current == "C":
                cat_c.append(entry)
            elif current == "D":
                cat_d.append(entry)
            else:
                cat_b.append(entry)

    return cat_c, cat_d, cat_b

# ─── 驗證邏輯 ─────────────────────────────────────────────────────────────────

def validate(cases: list[dict], appts: list[dict],
             cat_c: list[dict], cat_d: list[dict], cat_b: list[dict]) -> dict:

    issues: dict[str, list] = defaultdict(list)

    # 建立快查表
    case_by_code: dict[str, dict] = {c["case_code"]: c for c in cases if c["case_code"]}
    all_known_codes: set[str] = set(case_by_code.keys()) | {c["code"] for c in cat_c + cat_d + cat_b}

    # ── 1. 個案資料驗證 ──────────────────────────────────────────────────────

    for c in cases:
        code = c["case_code"] or f"row{c['row']}"
        name = c["name"]
        label = f"{code} {name}"

        # 必填欄位
        if not name:
            issues["case_required"].append({"id": code, "name": "（空白）", "field": "個案姓名", "msg": "姓名為空"})
        if not c["therapist"]:
            issues["case_required"].append({"id": code, "name": name, "field": "接案心理師", "msg": "心理師為空"})
        elif c["therapist"] not in THERAPIST_MAP:
            issues["case_therapist"].append({
                "id": code, "name": name,
                "therapist": c["therapist"],
                "msg": f"心理師「{c['therapist']}」不在系統心理師清單中，請確認名稱是否正確",
            })

        # 資料缺失（建議填）
        missing_fields = []
        if not c["national_id"]:
            missing_fields.append("身分證字號")
        if not c["birth_raw"]:
            missing_fields.append("出生日期")
        if not c["phone"] and not c["phone_home"]:
            missing_fields.append("連絡電話（手機/市話皆空）")
        if not c["gender_raw"] and c["gender_raw"] != 0:
            missing_fields.append("性別")
        if missing_fields:
            issues["case_missing"].append({
                "id": code, "name": name,
                "therapist": c["therapist"],
                "fields": missing_fields,
                "activatable": not (not c["national_id"] or not c["birth_raw"]),
                "msg": "缺少欄位：" + "、".join(missing_fields),
            })

        # 日期格式
        _, visit_err = parse_date(c["visit_raw"])
        if visit_err:
            issues["case_date"].append({"id": code, "name": name, "field": "進案日", "value": str(c["visit_raw"]), "msg": visit_err})
        _, birth_err = parse_date(c["birth_raw"])
        if birth_err:
            issues["case_date"].append({"id": code, "name": name, "field": "出生日期", "value": str(c["birth_raw"]), "msg": birth_err})

        # 身分證格式
        if c["national_id"] and not TW_ID_RE.match(c["national_id"].upper()):
            issues["case_id_format"].append({
                "id": code, "name": name,
                "value": c["national_id"],
                "msg": f"格式不符（應為 1 英文字母 + 數字1或2 + 8位數字），實際：{c['national_id']!r}",
            })

        # 可否啟用（拿到 8 碼編號）
        c["_activatable"] = bool(c["national_id"] and c["birth_raw"] and c["therapist"] in THERAPIST_MAP)
        c["_category"] = "A" if "115" in c["intake_year"] else "B"

    # ── 2. 伴侶案驗證 ────────────────────────────────────────────────────────

    couple_cases = [c for c in cases if is_couple_name(c["name"])]

    # 伴侶案應該兩兩配對（相同心理師）
    by_therapist: dict[str, list] = defaultdict(list)
    for c in couple_cases:
        by_therapist[c["therapist"]].append(c)

    for therapist, grp in by_therapist.items():
        if len(grp) % 2 != 0:
            issues["couple_odd"].append({
                "therapist": therapist,
                "count": len(grp),
                "cases": [{"id": c["case_code"], "name": c["name"]} for c in grp],
                "msg": f"心理師「{therapist}」有 {len(grp)} 個伴侶案（應為偶數），可能缺少配對的伴侶",
            })

    # 預約中的伴侶/合療 → 每個人都必須在 Sheet3 有自己的個案
    sheet3_names: set[str] = {
        c["name"].replace("(伴侶)", "").replace("（伴侶）", "").strip()
        for c in cases
    }
    for a in appts:
        if not (is_couple_name(a["case_name_raw"]) or is_couple_name(a["std_name"])):
            continue
        std = a["std_name"] or a["case_name_raw"]
        # 以「、」分割取出每個人名
        people = [p.strip() for p in std.split("、") if p.strip()]
        for pname in people:
            if pname and pname not in sheet3_names:
                # 避免重複
                already = any(
                    r["name"] == pname for r in issues["couple_missing_case"]
                )
                if not already:
                    issues["couple_missing_case"].append({
                        "name": pname,
                        "appt_sample": a["case_name_raw"],
                        "therapist": a["therapist"],
                        "msg": f"伴侶/合療預約中出現「{pname}」，但在個案清單（Sheet3）找不到此人的個案記錄",
                    })

    # ── 3. 預約 ↔ 個案交叉比對 ──────────────────────────────────────────────

    # 3a. 預約個案編號找不到個案資料
    appt_codes_with_issues: dict[str, list] = defaultdict(list)
    for a in appts:
        if a["case_code"] not in all_known_codes and not is_couple_name(a["case_name_raw"]) and "、" not in a["case_code"]:
            appt_codes_with_issues[a["case_code"]].append(a)

    for code, appt_list in appt_codes_with_issues.items():
        issues["appt_unknown_case"].append({
            "case_code": code,
            "sample_name": appt_list[0]["case_name_raw"],
            "count": len(appt_list),
            "therapist": appt_list[0]["therapist"],
            "msg": f"個案編號 {code!r}（{appt_list[0]['case_name_raw']}）在個案清單中找不到，共 {len(appt_list)} 筆預約",
        })

    # 3b. 心理師不一致（Sheet5 vs Sheet3 同一個案）
    case_therapist_in_s3 = {c["case_code"]: c["therapist"] for c in cases if c["case_code"]}
    mismatch_seen: set[str] = set()
    for a in appts:
        code = a["case_code"]
        if code in case_therapist_in_s3 and a["therapist"] and code not in mismatch_seen:
            if a["therapist"] != case_therapist_in_s3[code]:
                mismatch_seen.add(code)
                issues["therapist_mismatch"].append({
                    "case_code": code,
                    "case_name": case_by_code.get(code, {}).get("name", ""),
                    "s3_therapist": case_therapist_in_s3[code],
                    "appt_therapist": a["therapist"],
                    "msg": f"個案 {code} 在個案清單登記的心理師是「{case_therapist_in_s3[code]}」，但預約記錄的心理師是「{a['therapist']}」",
                })

    # 3c. 個案在清單但 2026 無預約
    appt_case_codes: set[str] = {a["case_code"] for a in appts}
    for c in cases:
        if c["case_code"] and c["case_code"] not in appt_case_codes:
            issues["case_no_appt"].append({
                "id": c["case_code"], "name": c["name"],
                "therapist": c["therapist"],
                "msg": f"個案 {c['case_code']}（{c['name']}）在個案清單中，但 2026 年無預約記錄",
            })

    # ── 4. 預約格式驗證 ──────────────────────────────────────────────────────

    for a in appts:
        # 未知心理師
        if a["therapist"] and a["therapist"] not in THERAPIST_MAP:
            issues["appt_therapist"].append({
                "row": a["row"], "date": a["date_raw"],
                "therapist": a["therapist"], "case_code": a["case_code"],
                "msg": f"心理師「{a['therapist']}」不在系統清單中",
            })
        # 未知場地
        if a["room"] and a["room"] not in ROOM_CODES:
            issues["appt_room"].append({
                "row": a["row"], "date": a["date_raw"],
                "room": a["room"], "case_code": a["case_code"],
                "msg": f"場地代碼「{a['room']}」不在系統清單中（需新增房間）",
            })

    # ── 5. 去重：couple_missing_case 與 appt_unknown_case 重疊時，前者不顯示 ──
    # appt_unknown_case 的 sample_name 是 case_name_raw（solo 預約即為個人姓名）
    unknown_case_names: set[str] = {r["sample_name"] for r in issues["appt_unknown_case"]}
    # 另外也收集 std_name 拆出的個別名字，避免名稱不一致漏掉
    for r in issues["appt_unknown_case"]:
        for part in r["sample_name"].split("、"):
            unknown_case_names.add(part.strip())
    issues["couple_missing_case"] = [
        r for r in issues["couple_missing_case"]
        if r["name"] not in unknown_case_names
    ]

    # ── 6. 統計摘要 ──────────────────────────────────────────────────────────

    activatable = sum(1 for c in cases if c.get("_activatable"))
    cat_counts = {"A": sum(1 for c in cases if c.get("_category") == "A"),
                  "B": sum(1 for c in cases if c.get("_category") == "B")}

    summary = {
        "total_cases": len(cases),
        "activatable": activatable,
        "cat_A": cat_counts["A"],
        "cat_B": cat_counts["B"],
        "cat_C": len(cat_c),
        "cat_D": len(cat_d),
        "total_appts": len(appts),
        "issue_counts": {k: len(v) for k, v in issues.items()},
        "total_issues": sum(len(v) for v in issues.values()),
    }

    return {"summary": summary, "issues": dict(issues), "cat_c": cat_c, "cat_d": cat_d}

# ─── FastAPI App ──────────────────────────────────────────────────────────────

app = FastAPI(title="慈恩資料稽核工具")

HTML_PAGE = r"""<!DOCTYPE html>
<html lang="zh-TW">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>慈恩資料遷移稽核工具</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:-apple-system,"PingFang TC",sans-serif;font-size:14px;background:#f3f4f6;color:#111}
header{background:#1e3a5f;color:white;padding:18px 28px;display:flex;align-items:center;gap:12px}
header h1{font-size:18px;font-weight:700}
header .sub{font-size:13px;opacity:.7;margin-top:2px}

/* Upload zone */
.upload-area{margin:24px 28px;background:white;border-radius:12px;padding:32px;
  border:2px dashed #cbd5e1;text-align:center;cursor:pointer;transition:all .2s;
  box-shadow:0 1px 3px rgba(0,0,0,.05)}
.upload-area.drag{border-color:#1e3a5f;background:#f0f4ff}
.upload-area.loading{border-color:#6b7280;cursor:default}
.upload-icon{font-size:36px;margin-bottom:12px}
.upload-title{font-size:16px;font-weight:600;color:#1e3a5f;margin-bottom:6px}
.upload-sub{font-size:13px;color:#6b7280}
.upload-file-name{font-size:13px;color:#4b5563;margin-top:10px;font-weight:500}
.spinner{display:none;width:28px;height:28px;border:3px solid #e5e7eb;
  border-top-color:#1e3a5f;border-radius:50%;animation:spin .8s linear infinite;margin:0 auto 12px}
@keyframes spin{to{transform:rotate(360deg)}}

/* Summary bar */
.summary{display:flex;flex-wrap:wrap;gap:10px;margin:0 28px 20px}
.s-card{background:white;border-radius:8px;padding:14px 18px;min-width:120px;
  box-shadow:0 1px 3px rgba(0,0,0,.06);flex:1}
.s-card .lbl{font-size:11px;color:#9ca3af;margin-bottom:4px}
.s-card .val{font-size:22px;font-weight:700}
.s-card.red .val{color:#dc2626}
.s-card.amber .val{color:#d97706}
.s-card.green .val{color:#16a34a}
.s-card.blue .val{color:#2563eb}

/* Sections */
.sections{margin:0 28px 40px;display:flex;flex-direction:column;gap:12px}
.section{background:white;border-radius:10px;box-shadow:0 1px 3px rgba(0,0,0,.06);overflow:hidden}
.sec-header{display:flex;align-items:center;gap:10px;padding:14px 18px;cursor:pointer;
  border-left:4px solid transparent;user-select:none}
.sec-header:hover{background:#f9fafb}
.sec-header.error{border-left-color:#dc2626}
.sec-header.warning{border-left-color:#d97706}
.sec-header.info{border-left-color:#2563eb}
.sec-header.good{border-left-color:#16a34a}
.sec-title{font-size:14px;font-weight:600;flex:1}
.sec-count{font-size:12px;padding:2px 10px;border-radius:10px;font-weight:600}
.badge-red{background:#fee2e2;color:#b91c1c}
.badge-amber{background:#fef3c7;color:#92400e}
.badge-green{background:#dcfce7;color:#166534}
.badge-blue{background:#dbeafe;color:#1e40af}
.sec-arrow{color:#9ca3af;font-size:12px;transition:transform .2s}
.sec-content{display:none;padding:0 18px 16px;border-top:1px solid #f0f0f0}
.sec-content.open{display:block}

/* Tables */
.tbl-wrap{overflow-x:auto;margin-top:12px}
table{width:100%;border-collapse:collapse;font-size:13px}
th{background:#f8fafc;padding:8px 10px;text-align:left;font-weight:600;
   border-bottom:2px solid #e5e7eb;white-space:nowrap;color:#374151}
td{padding:8px 10px;border-bottom:1px solid #f3f4f6;vertical-align:top}
tr:hover td{background:#fafafa}
code{background:#f1f5f9;padding:1px 5px;border-radius:3px;font-size:12px;font-family:monospace}
.tag{display:inline-block;padding:1px 7px;border-radius:8px;font-size:11px;font-weight:600;margin:1px}
.tag-red{background:#fee2e2;color:#b91c1c}
.tag-amber{background:#fef3c7;color:#92400e}
.tag-blue{background:#dbeafe;color:#1e40af}
.tag-purple{background:#ede9fe;color:#5b21b6}
.tag-gray{background:#f3f4f6;color:#374151}

/* Filters */
.filters{display:flex;gap:8px;flex-wrap:wrap;align-items:center;margin-top:12px}
.filters input,.filters select{padding:5px 10px;border:1px solid #d1d5db;border-radius:6px;
  font-size:13px;min-width:160px}
.filters input:focus,.filters select:focus{outline:none;border-color:#1e3a5f}

/* Empty state */
.empty{text-align:center;padding:32px;color:#9ca3af;font-size:13px}

/* No file state */
#no-file{text-align:center;padding:40px;color:#9ca3af;font-size:14px}
#results{display:none}
</style>
</head>
<body>
<header>
  <div>
    <h1>🏥 慈恩資料遷移稽核工具</h1>
    <div class="sub">上傳 xlsx → 驗證 → 修正 → 重新上傳</div>
  </div>
</header>

<!-- Upload Zone -->
<div class="upload-area" id="drop-zone" ondragover="onDragOver(event)" ondragleave="onDragLeave()"
     ondrop="onDrop(event)" onclick="document.getElementById('file-input').click()">
  <div class="spinner" id="spinner"></div>
  <div class="upload-icon" id="upload-icon">📂</div>
  <div class="upload-title" id="upload-title">拖曳 xlsx 檔案到這裡，或點擊選擇</div>
  <div class="upload-sub">支援 .xlsx 格式　每次上傳後自動重新驗證</div>
  <div class="upload-file-name" id="file-name"></div>
</div>
<input type="file" id="file-input" accept=".xlsx" style="display:none" onchange="onFileSelected(this)">

<div id="no-file"><p>還沒有上傳檔案</p></div>

<!-- Results -->
<div id="results">
  <div class="summary" id="summary-bar"></div>
  <div class="sections" id="sections"></div>
</div>

<script>
let lastFile = null;

function onDragOver(e){e.preventDefault();document.getElementById('drop-zone').classList.add('drag')}
function onDragLeave(){document.getElementById('drop-zone').classList.remove('drag')}
function onDrop(e){
  e.preventDefault();
  document.getElementById('drop-zone').classList.remove('drag');
  const f = e.dataTransfer.files[0];
  if(f) uploadFile(f);
}
function onFileSelected(input){
  if(input.files[0]) uploadFile(input.files[0]);
}

function setLoading(loading, fileName){
  document.getElementById('spinner').style.display = loading ? 'block' : 'none';
  document.getElementById('upload-icon').style.display = loading ? 'none' : 'block';
  document.getElementById('upload-title').textContent = loading ? '驗證中…' : '拖曳 xlsx 檔案到這裡，或點擊選擇';
  document.getElementById('drop-zone').classList.toggle('loading', loading);
  if(fileName) document.getElementById('file-name').textContent = '📄 ' + fileName;
}

async function uploadFile(file){
  if(!file.name.endsWith('.xlsx')){alert('請選擇 .xlsx 格式的檔案'); return;}
  lastFile = file;
  setLoading(true, file.name);
  document.getElementById('no-file').style.display = 'none';
  document.getElementById('results').style.display = 'none';

  const fd = new FormData();
  fd.append('file', file);
  try{
    const resp = await fetch('/validate', {method:'POST', body:fd});
    if(!resp.ok){const t=await resp.text(); throw new Error(t);}
    const data = await resp.json();
    setLoading(false, file.name);
    renderResults(data);
  } catch(e){
    setLoading(false, file.name);
    alert('驗證失敗：' + e.message);
  }
}

function renderResults(data){
  const {summary, issues, cat_c, cat_d} = data;
  document.getElementById('results').style.display = 'block';
  renderSummary(summary);
  renderSections(issues, cat_c, cat_d, summary);
}

function renderSummary(s){
  const bar = document.getElementById('summary-bar');
  const totalIssues = s.total_issues;
  bar.innerHTML = `
    <div class="s-card ${totalIssues===0?'green':'red'}">
      <div class="lbl">總問題數</div>
      <div class="val">${totalIssues===0?'✓ 無問題':totalIssues}</div>
    </div>
    <div class="s-card blue">
      <div class="lbl">個案總數</div>
      <div class="val">${s.total_cases}</div>
    </div>
    <div class="s-card green">
      <div class="lbl">可啟用（可拿8碼）</div>
      <div class="val">${s.activatable}</div>
    </div>
    <div class="s-card blue">
      <div class="lbl">預約總數</div>
      <div class="val">${s.total_appts}</div>
    </div>
    <div class="s-card amber">
      <div class="lbl">Category C（未建檔≥3）</div>
      <div class="val">${s.cat_C}</div>
    </div>
    <div class="s-card amber">
      <div class="lbl">Category D（未建檔<3）</div>
      <div class="val">${s.cat_D}</div>
    </div>`;
}

const SECTION_DEFS = [
  {key:'case_required', title:'❗ 必填欄位缺失', cls:'error', badge:'red',
    desc:'個案缺少姓名或心理師，無法匯入'},
  {key:'case_therapist', title:'⚠️ 心理師名稱找不到對應', cls:'warning', badge:'amber',
    desc:'心理師姓名不在系統清單，請確認名稱是否有誤或需新增心理師帳號'},
  {key:'case_missing', title:'📋 個案資料不完整', cls:'warning', badge:'amber',
    desc:'缺少建議欄位（身分證、出生日期、電話、性別）。有身分證+出生日期才能啟用取得8碼編號'},
  {key:'case_date', title:'📅 日期格式錯誤', cls:'error', badge:'red',
    desc:'日期欄位無法解析，請修正格式（民國：82/07/12，或西元：1993/07/12）'},
  {key:'case_id_format', title:'🪪 身分證格式不符', cls:'warning', badge:'amber',
    desc:'台灣身分證應為：1英文 + 數字1或2 + 8位數字（共10碼）'},
  {key:'couple_odd', title:'👫 伴侶案配對不完整', cls:'error', badge:'red',
    desc:'有「伴侶」的個案應兩兩配對（相同心理師），若為奇數則有缺失'},
  {key:'couple_missing_case', title:'👫 預約伴侶在個案清單中找不到', cls:'warning', badge:'amber',
    desc:'預約記錄中出現的伴侶/合併名稱，在個案清單（Sheet3）中找不到對應個案'},
  {key:'appt_unknown_case', title:'🔗 預約找不到對應個案', cls:'warning', badge:'amber',
    desc:'Sheet5 預約的個案編號，在個案清單和未建檔清單中都找不到'},
  {key:'therapist_mismatch', title:'⚡ 心理師不一致（個案 vs 預約）', cls:'warning', badge:'amber',
    desc:'同一個案在個案清單和預約記錄中的心理師不同'},
  {key:'case_no_appt', title:'📭 個案清單中有但無預約', cls:'info', badge:'blue',
    desc:'個案有基本資料，但 2026 年無預約記錄'},
  {key:'appt_therapist', title:'⚠️ 預約心理師找不到對應', cls:'warning', badge:'amber',
    desc:'預約記錄的心理師名稱不在系統清單中'},
  {key:'appt_room', title:'🚪 未知場地代碼', cls:'info', badge:'blue',
    desc:'場地代碼不在系統房間清單中（需先新增該房間）'},
];

function renderSections(issues, cat_c, cat_d, summary){
  const container = document.getElementById('sections');
  container.innerHTML = '';

  for(const def of SECTION_DEFS){
    const rows = issues[def.key] || [];
    const count = rows.length;
    const sec = document.createElement('div');
    sec.className = 'section';
    const isEmpty = count === 0;
    sec.innerHTML = `
      <div class="sec-header ${isEmpty?'good':def.cls}" onclick="toggleSection(this)">
        <span class="sec-title">${def.title}</span>
        <span class="sec-count badge-${isEmpty?'green':def.badge}">${isEmpty?'✓ 無問題':count+' 個問題'}</span>
        <span class="sec-arrow">▼</span>
      </div>
      <div class="sec-content ${!isEmpty?'open':''}">
        <p style="font-size:12px;color:#6b7280;margin-top:12px">${def.desc}</p>
        ${isEmpty ? '<div class="empty">✓ 此項目無問題</div>' : renderTable(def.key, rows)}
      </div>`;
    container.appendChild(sec);
  }

  // Category C/D section
  if(cat_c.length || cat_d.length){
    const sec = document.createElement('div');
    sec.className = 'section';
    sec.innerHTML = `
      <div class="sec-header warning" onclick="toggleSection(this)">
        <span class="sec-title">🆕 未建檔個案清單（需補建資料）</span>
        <span class="sec-count badge-amber">C類: ${cat_c.length} 人 / D類: ${cat_d.length} 人</span>
        <span class="sec-arrow">▼</span>
      </div>
      <div class="sec-content">
        <h4 style="margin:12px 0 6px;font-size:13px">❗ Category C — 預約 ≥ 3 次（優先補建）</h4>
        <div class="tbl-wrap"><table>
          <thead><tr><th>代號</th><th>姓名</th><th>2026預約次數</th><th>心理師</th><th>建議行動</th></tr></thead>
          <tbody>${cat_c.map(r=>`<tr><td><code>${r.code}</code></td><td>${r.name}</td><td>${r.appt_count}</td><td>${r.therapist}</td><td>${r.note}</td></tr>`).join('')}</tbody>
        </table></div>
        <h4 style="margin:16px 0 6px;font-size:13px">⚪ Category D — 預約 &lt; 3 次（低優先）</h4>
        <div class="tbl-wrap"><table>
          <thead><tr><th>代號</th><th>姓名</th><th>2026預約次數</th><th>心理師</th><th>備註</th></tr></thead>
          <tbody>${cat_d.map(r=>`<tr><td><code>${r.code}</code></td><td>${r.name}</td><td>${r.appt_count}</td><td>${r.therapist}</td><td>${r.note}</td></tr>`).join('')}</tbody>
        </table></div>
      </div>`;
    container.appendChild(sec);
  }
}

function renderTable(key, rows){
  if(!rows.length) return '<div class="empty">無資料</div>';

  const cols = {
    case_required: ['id','name','field','msg'],
    case_therapist: ['id','name','therapist','msg'],
    case_missing: ['id','name','therapist','fields','activatable','msg'],
    case_date: ['id','name','field','value','msg'],
    case_id_format: ['id','name','value','msg'],
    couple_odd: ['therapist','count','cases','msg'],
    couple_missing_case: ['name','therapist','appt_sample','msg'],
    appt_unknown_case: ['case_code','sample_name','therapist','count','msg'],
    therapist_mismatch: ['case_code','case_name','s3_therapist','appt_therapist','msg'],
    case_no_appt: ['id','name','therapist','msg'],
    appt_therapist: ['row','date','therapist','case_code','msg'],
    appt_room: ['row','date','room','case_code','msg'],
  };

  const headers = {
    id:'個案編號', name:'姓名', field:'欄位', msg:'問題說明', therapist:'心理師',
    value:'原始值', fields:'缺少欄位', activatable:'可啟用8碼?',
    count:'筆數', cases:'相關個案', case_code:'個案編號', sample_name:'個案名稱',
    s3_therapist:'個案清單心理師', appt_therapist:'預約心理師',
    case_name:'個案姓名', row:'行號', date:'日期', room:'場地',
    appt_sample:'預約原始名稱',
  };

  const keys = cols[key] || Object.keys(rows[0]).filter(k=>k!=='row'||key.startsWith('appt'));

  let html = '<div class="tbl-wrap"><table><thead><tr>';
  for(const k of keys) html += `<th>${headers[k]||k}</th>`;
  html += '</tr></thead><tbody>';

  for(const r of rows){
    html += '<tr>';
    for(const k of keys){
      let v = r[k];
      if(k==='fields' && Array.isArray(v)){
        v = v.map(f=>`<span class="tag tag-amber">${f}</span>`).join('');
      } else if(k==='cases' && Array.isArray(v)){
        v = v.map(c=>`<span class="tag tag-red"><code>${c.id}</code> ${c.name}</span>`).join('');
      } else if(k==='activatable'){
        v = v ? '<span class="tag tag-green">✓ 可啟用</span>' : '<span class="tag tag-red">✗ 缺身分證或出生日</span>';
      } else if(k==='case_code'||k==='id'){
        v = `<code>${v||'—'}</code>`;
      } else if(k==='msg'){
        v = `<span style="color:#374151">${v||''}</span>`;
      } else {
        v = v === null || v === undefined ? '—' : String(v);
      }
      html += `<td>${v}</td>`;
    }
    html += '</tr>';
  }
  html += '</tbody></table></div>';
  return html;
}

function toggleSection(header){
  const content = header.nextElementSibling;
  const arrow = header.querySelector('.sec-arrow');
  const isOpen = content.classList.contains('open');
  content.classList.toggle('open', !isOpen);
  arrow.style.transform = isOpen ? '' : 'rotate(180deg)';
}
</script>
</body>
</html>"""

@app.get("/", response_class=HTMLResponse)
async def index():
    return HTMLResponse(content=HTML_PAGE)

@app.post("/validate")
async def validate_file(file: UploadFile = File(...)):
    if not file.filename.endswith(".xlsx"):
        return JSONResponse(status_code=400, content={"error": "請上傳 .xlsx 格式的檔案"})
    try:
        data = await file.read()
        wb = load_workbook_from_bytes(data)
        cases = read_sheet3(wb)
        appts, skipped = read_sheet5(wb)
        cat_c, cat_d, cat_b = read_problem_cases(wb)
        result = validate(cases, appts, cat_c, cat_d, cat_b)
        result["skipped"] = skipped
        return JSONResponse(content=result)
    except KeyError as e:
        return JSONResponse(status_code=422, content={"error": f"找不到工作表：{e}，請確認上傳的是正確的 xlsx 檔案"})
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})

if __name__ == "__main__":
    print("=" * 50)
    print("  慈恩資料遷移稽核工具")
    print("  http://localhost:8787")
    print("  Ctrl+C 停止")
    print("=" * 50)
    uvicorn.run(app, host="0.0.0.0", port=8787, log_level="warning")
