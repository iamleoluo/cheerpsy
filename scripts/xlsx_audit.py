#!/usr/bin/env python3.11
"""
xlsx_audit.py — 慈恩資料遷移稽核工具

讀取 2026年個案資料稽核報告_v2.xlsx，驗證每筆資料是否符合 CheerPsy DB 格式，
輸出可在瀏覽器直接開啟的互動式 HTML 稽核報告。

用法:
    python3.11 scripts/xlsx_audit.py
    python3.11 scripts/xlsx_audit.py --xlsx /path/to/file.xlsx --output output/report.html
"""

import argparse
import json
import re
import sys
from datetime import date, datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("請先安裝 openpyxl: python3.11 -m pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# ─── 靜態映射表（不連 DB） ────────────────────────────────────────────────────

THERAPIST_MAP: dict[str, str] = {
    "呂孟育": "T001", "林紀宇": "T002", "林容蒂": "T003",
    "邱似齡": "T004", "蔡孟潔": "T005", "陳慧苓": "T006",
    "游子瑩": "T007", "葉邦彥": "T008", "鄭幼毅": "T009",
    "楊顯欽": "T010", "劉柏宏": "T011", "邱惟雅": "T012",
    "潘柔靜": "T013", "邱意祺": "T014", "劉彥君": "T015",
    "羅紀萱": "T016", "黃慧婷": "T017",
}

ROOM_CODES: set[str] = {"1A", "1B", "1C", "1D", "2A", "2B", "2C", "2D", "2E", "3A", "3B", "3C", "3D", "3E", "3F"}

SELF_PAY_SOURCES: set[str] = {
    "自費", "自費伴侶", "自費(個別轉伴侶)", "自費(家族)", "家事自費",
    "教支轉自費", "澎湖教支轉自費",
}

INSTITUTION_MAP: dict[str, str] = {
    "教支": "教育支持方案", "台南市教支": "教育支持方案",
    "台積電": "台積電EAP", "台積EAP": "台積電EAP", "台積EAP伴侶": "台積電EAP", "蛹之生台積電": "台積電EAP",
    "台電": "台電", "奇美": "奇美",
    "衛生局": "衛生局",
    "台南市家暴性侵防治中心": "家暴性侵防治中心",
    "台南市政府人事處": "人事處", "人事處": "人事處",
    "南家扶": "家扶基金會", "家扶FAP": "家扶基金會",
    "善牧": "善牧基金會",
    "國軍": "國軍",
    "地方法院": "法院", "法院": "法院", "法院EAP": "法院",
    "女權家事": "女權", "家事(女權)": "女權",
    "更生": "更生",
    "職災中心": "職災",
    "警局": "警局", "台南市警局": "警局",
    "青壯": "青壯方案",
    "脆家": "脆弱家庭",
    "醫事": "醫事人員方案", "醫事人員方案": "醫事人員方案",
    "家防": "家庭防暴", "家防伴侶": "家庭防暴",
    "永仁": "永仁",
    "永觀": "永觀", "永觀長照": "永觀",
    "戀戀交友": "戀戀交友",
    "荷光轉案伴侶": "荷光",
    "伴侶招募": "?待確認", "伴家加久方案": "?待確認",
    "實習生方案": "?待確認", "15-30方案": "?待確認", "15-45": "?待確認",
}

TW_ID_RE = re.compile(r'^[A-Z][12]\d{8}$')
DATE_RE = re.compile(r'^(\d{2,4})[/.](\d{1,2})[/.](\d{1,2})$')

# ─── 日期解析 ─────────────────────────────────────────────────────────────────

def parse_date(raw) -> tuple[date | None, str | None]:
    """Returns (date, error_message). date is None on failure."""
    if raw is None or raw == "":
        return None, None
    if isinstance(raw, (datetime, date)):
        return raw.date() if isinstance(raw, datetime) else raw, None
    s = str(raw).strip()
    m = DATE_RE.match(s)
    if not m:
        return None, f"無法解析日期格式：{s!r}"
    y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
    if y < 200:
        y += 1911
    try:
        return date(y, mo, d), None
    except ValueError as e:
        return None, f"日期值無效 {y}-{mo:02d}-{d:02d}：{e}"

# ─── 性別正規化 ───────────────────────────────────────────────────────────────

_FULLWIDTH = str.maketrans("０１２３４５６７８９", "0123456789")

def parse_gender(raw) -> tuple[str | None, str | None]:
    if raw is None or raw == "":
        return None, None
    s = str(raw).strip().translate(_FULLWIDTH)
    if s in ("1", "男"):
        return "male", None
    if s in ("2", "女"):
        return "female", None
    return "other", f"未知性別值：{raw!r}，已設為 other"

# ─── 身分證驗證 ───────────────────────────────────────────────────────────────

def validate_tw_id(raw) -> str | None:
    if not raw:
        return None
    s = str(raw).strip().upper()
    if not TW_ID_RE.match(s):
        return f"身分證格式不符：{s!r}（應為 1 英文 + 1 位(1或2) + 8 數字）"
    return None

# ─── 讀取 Sheet3 個案資料 ─────────────────────────────────────────────────────

def read_cases(ws) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(c).strip() if c else "" for c in rows[0]]

    def col(row, name):
        try:
            return row[header.index(name)]
        except ValueError:
            return None

    cases = []
    for i, row in enumerate(rows[1:], start=2):
        if not any(row):
            continue
        issues: list[dict] = []

        case_code = str(col(row, "個案編號") or "").strip()
        name = str(col(row, "個案姓名") or "").strip()
        therapist_raw = str(col(row, "接案心理師") or "").strip()
        funding_raw = str(col(row, "經費來源") or "").strip()
        gender_raw = col(row, "性別")
        national_id_raw = col(row, "身分證字號")
        birth_raw = col(row, "出生日期")
        visit_raw = col(row, "進案日")
        intake_year = str(col(row, "個案進案年份") or "").strip()
        appt_count = col(row, "2026預約次數") or 0

        # 分類
        category = "A" if "115" in intake_year else "B"

        # 個案編號（舊系統流水號，僅做參考，格式不強制）
        if not case_code:
            issues.append({"level": "warning", "field": "個案編號", "msg": "個案編號為空（匯入後靠 temp_seq 識別）"})

        # 姓名
        if not name:
            issues.append({"level": "error", "field": "個案姓名", "msg": "姓名為空"})
        elif len(name) > 100:
            issues.append({"level": "error", "field": "個案姓名", "msg": f"姓名超過 100 字元（{len(name)}）"})

        # 心理師
        therapist_code = THERAPIST_MAP.get(therapist_raw)
        if not therapist_raw:
            issues.append({"level": "error", "field": "接案心理師", "msg": "心理師為空"})
        elif not therapist_code:
            issues.append({"level": "error", "field": "接案心理師", "msg": f"心理師名稱找不到對應：{therapist_raw!r}"})

        # 進案日
        visit_date, visit_err = parse_date(visit_raw)
        if visit_err:
            issues.append({"level": "error", "field": "進案日", "msg": visit_err})

        # 出生日期
        birth_date, birth_err = parse_date(birth_raw)
        if birth_err:
            issues.append({"level": "warning", "field": "出生日期", "msg": birth_err})

        # 性別
        gender, gender_warn = parse_gender(gender_raw)
        if gender_warn:
            issues.append({"level": "warning", "field": "性別", "msg": gender_warn})

        # 身分證
        id_warn = validate_tw_id(national_id_raw)
        if id_warn:
            issues.append({"level": "warning", "field": "身分證字號", "msg": id_warn})

        # 經費來源
        funding_source = None
        institution = None
        fs = funding_raw.strip() if funding_raw else ""
        if fs in SELF_PAY_SOURCES or not fs:
            funding_source = "self_pay"
        elif fs in INSTITUTION_MAP:
            funding_source = "institution"
            institution = INSTITUTION_MAP[fs]
            if institution.startswith("?"):
                issues.append({"level": "warning", "field": "經費來源", "msg": f"機構名稱待確認：{fs!r}"})
        else:
            funding_source = "institution"
            institution = f"?{fs}"
            issues.append({"level": "warning", "field": "經費來源", "msg": f"未知經費來源，需確認機構：{fs!r}"})

        # 整體狀態
        has_error = any(i["level"] == "error" for i in issues)
        has_warning = any(i["level"] == "warning" for i in issues)
        status = "error" if has_error else ("warning" if has_warning else "ok")

        cases.append({
            "row": i,
            "case_code": case_code,
            "name": name,
            "therapist": therapist_raw,
            "therapist_code": therapist_code,
            "category": category,
            "intake_year": intake_year,
            "appt_count": int(appt_count) if appt_count else 0,
            "visit_date": str(visit_date) if visit_date else str(visit_raw or ""),
            "birth_date": str(birth_date) if birth_date else str(birth_raw or ""),
            "gender": gender,
            "national_id": str(national_id_raw or ""),
            "funding_source": funding_source,
            "institution": institution,
            "funding_raw": funding_raw,
            "status": status,
            "issues": issues,
        })

    return cases

# ─── 讀取 Sheet2 未處理個案 ───────────────────────────────────────────────────

def read_problem_cases(ws) -> tuple[list[dict], list[dict], list[dict]]:
    """Returns (category_c, category_d, category_b)"""
    rows = list(ws.iter_rows(values_only=True))
    cat_c, cat_d, cat_b = [], [], []
    current = None
    current_header = None

    for i, row in enumerate(rows, start=1):
        cell1 = str(row[1] or "").strip()
        if "完全未列入" in cell1 and "≥3" in cell1:
            current = "C"
            current_header = None
        elif "完全未列入" in cell1 and "<3" in cell1:
            current = "D"
            current_header = None
        elif "跨年延續" in cell1:
            current = "B"
            current_header = None
        elif cell1 == "個案姓名":
            current_header = True
        elif current and current_header and row[0]:
            code = str(row[0]).strip()
            entry = {
                "code": code,
                "name": str(row[1] or "").strip(),
                "appt_count": row[2] or 0,
                "therapist": str(row[3] or "").strip(),
                "note": str(row[4] or "").strip() if len(row) > 4 else "",
                "category": current,
            }
            if current == "C":
                cat_c.append(entry)
            elif current == "D":
                cat_d.append(entry)
            elif current == "B":
                cat_b.append(entry)

    return cat_c, cat_d, cat_b

# ─── 讀取 Sheet5 預約資料 ─────────────────────────────────────────────────────

def read_appointments(ws, case_codes: set[str]) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    DATE_STRIP_RE = re.compile(r'[（(][^）)]+[）)]')
    TIME_RE = re.compile(r'^\d{1,2}:\d{2}$')

    appts = []
    skipped = 0

    for i, row in enumerate(rows[1:], start=2):
        if not any(row):
            continue

        # 欄位: 日期, 開始, 結束, 場地, 心理師, 個案, 標準化姓名, 個案編號, 備註, 建立者, 週期
        raw_date = str(row[0] or "").strip()
        raw_start = str(row[1] or "").strip()
        raw_end = str(row[2] or "").strip()
        room_code = str(row[3] or "").strip()
        therapist_raw = str(row[4] or "").strip()
        case_name = str(row[5] or "").strip()
        std_name = str(row[6] or "").strip() if len(row) > 6 else ""
        case_code = str(row[7] or "").strip() if len(row) > 7 else ""
        notes = str(row[8] or "").strip() if len(row) > 8 else ""

        # 無個案編號 → 跳過
        if not case_code:
            skipped += 1
            continue

        issues: list[dict] = []

        # 日期解析
        date_str = DATE_STRIP_RE.sub("", raw_date).strip()
        parsed_date = None
        if date_str:
            try:
                parsed_date = datetime.strptime(date_str, "%Y/%m/%d").date()
            except ValueError:
                issues.append({"level": "error", "field": "日期", "msg": f"無法解析日期：{raw_date!r}"})
        else:
            issues.append({"level": "error", "field": "日期", "msg": "日期為空"})

        # 時間
        if not TIME_RE.match(raw_start):
            issues.append({"level": "error", "field": "開始時間", "msg": f"格式錯誤：{raw_start!r}"})
        if not TIME_RE.match(raw_end):
            issues.append({"level": "error", "field": "結束時間", "msg": f"格式錯誤：{raw_end!r}"})

        # 場地
        if room_code and room_code not in ROOM_CODES:
            issues.append({"level": "warning", "field": "場地", "msg": f"未知場地代碼：{room_code!r}"})

        # 心理師
        therapist_code = THERAPIST_MAP.get(therapist_raw)
        if not therapist_raw:
            issues.append({"level": "error", "field": "心理師", "msg": "心理師為空"})
        elif not therapist_code:
            issues.append({"level": "warning", "field": "心理師", "msg": f"心理師名稱找不到對應：{therapist_raw!r}"})

        # 個案編號
        if case_code not in case_codes:
            issues.append({"level": "warning", "field": "個案編號", "msg": f"個案編號在個案清單中找不到：{case_code!r}"})

        has_error = any(i["level"] == "error" for i in issues)
        has_warning = any(i["level"] == "warning" for i in issues)
        status = "error" if has_error else ("warning" if has_warning else "ok")

        appts.append({
            "row": i,
            "date": str(parsed_date) if parsed_date else raw_date,
            "start": raw_start,
            "end": raw_end,
            "room": room_code,
            "therapist": therapist_raw,
            "therapist_code": therapist_code,
            "case_name": case_name,
            "std_name": std_name,
            "case_code": case_code,
            "notes": notes,
            "status": status,
            "issues": issues,
        })

    return appts, skipped

# ─── HTML 模板 ────────────────────────────────────────────────────────────────

HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="zh-TW">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>慈恩資料遷移稽核報告</title>
<style>
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body { font-family: -apple-system, "PingFang TC", sans-serif; font-size: 14px; background: #f5f5f5; color: #222; }
  h1 { padding: 20px 24px 8px; font-size: 20px; font-weight: 700; }
  .subtitle { padding: 0 24px 16px; color: #666; font-size: 13px; }

  /* Stats bar */
  .stats { display: flex; gap: 12px; padding: 0 24px 20px; flex-wrap: wrap; }
  .stat-card { background: white; border-radius: 8px; padding: 14px 20px; min-width: 140px;
               box-shadow: 0 1px 3px rgba(0,0,0,.1); }
  .stat-card .label { font-size: 12px; color: #888; margin-bottom: 4px; }
  .stat-card .value { font-size: 24px; font-weight: 700; }
  .stat-card.error .value { color: #d32f2f; }
  .stat-card.warning .value { color: #e65100; }
  .stat-card.ok .value { color: #2e7d32; }
  .stat-card.info .value { color: #1565c0; }

  /* Tabs */
  .tabs { display: flex; gap: 0; padding: 0 24px; border-bottom: 2px solid #ddd; margin-bottom: 0; }
  .tab { padding: 10px 20px; cursor: pointer; font-size: 14px; font-weight: 500;
         border-bottom: 3px solid transparent; margin-bottom: -2px; color: #666; }
  .tab.active { color: #1565c0; border-bottom-color: #1565c0; }

  /* Panels */
  .panel { display: none; padding: 20px 24px; }
  .panel.active { display: block; }

  /* Filters */
  .filters { display: flex; gap: 10px; margin-bottom: 16px; flex-wrap: wrap; align-items: center; }
  .filter-btn { padding: 5px 14px; border: 1px solid #ccc; border-radius: 16px; background: white;
               cursor: pointer; font-size: 13px; transition: all .15s; }
  .filter-btn.active { background: #1565c0; color: white; border-color: #1565c0; }
  .filter-btn.error.active { background: #d32f2f; border-color: #d32f2f; }
  .filter-btn.warning.active { background: #e65100; border-color: #e65100; }
  .filter-btn.ok.active { background: #2e7d32; border-color: #2e7d32; }
  select, input { padding: 5px 10px; border: 1px solid #ccc; border-radius: 6px; font-size: 13px; }

  /* Table */
  .table-wrap { overflow-x: auto; background: white; border-radius: 8px;
               box-shadow: 0 1px 3px rgba(0,0,0,.1); }
  table { width: 100%; border-collapse: collapse; }
  th { background: #f8f8f8; padding: 10px 12px; text-align: left; font-size: 13px;
       font-weight: 600; border-bottom: 2px solid #ddd; white-space: nowrap; }
  td { padding: 9px 12px; border-bottom: 1px solid #f0f0f0; font-size: 13px; vertical-align: top; }
  tr:hover td { background: #fafafa; }
  tr.error td { background: #fff5f5; }
  tr.warning td { background: #fffbf0; }
  tr.ok td { background: white; }

  /* Status badges */
  .badge { display: inline-block; padding: 2px 8px; border-radius: 10px; font-size: 11px; font-weight: 600; }
  .badge.error { background: #ffebee; color: #c62828; }
  .badge.warning { background: #fff3e0; color: #bf360c; }
  .badge.ok { background: #e8f5e9; color: #1b5e20; }
  .badge.cat-A { background: #e3f2fd; color: #0d47a1; }
  .badge.cat-B { background: #f3e5f5; color: #4a148c; }
  .badge.cat-C { background: #ffebee; color: #b71c1c; }
  .badge.cat-D { background: #fff8e1; color: #e65100; }

  /* Issues list */
  .issues { margin-top: 4px; }
  .issue { font-size: 11px; padding: 2px 0; }
  .issue.error { color: #c62828; }
  .issue.warning { color: #e65100; }

  /* Pagination */
  .pagination { display: flex; gap: 8px; padding: 16px 0; justify-content: center; flex-wrap: wrap; }
  .page-btn { padding: 5px 12px; border: 1px solid #ccc; border-radius: 4px; cursor: pointer;
             background: white; font-size: 13px; }
  .page-btn.active { background: #1565c0; color: white; border-color: #1565c0; }
  .page-btn:disabled { opacity: .4; cursor: default; }

  /* count info */
  .count-info { font-size: 13px; color: #666; margin-bottom: 12px; }

  /* Legend */
  .legend { font-size: 12px; color: #888; padding: 8px 0 0; }
</style>
</head>
<body>
<h1>慈恩資料遷移稽核報告</h1>
<div class="subtitle">來源：2026年個案資料稽核報告_v2.xlsx　製表：<span id="gen-time"></span></div>

<div class="stats" id="stats-bar"></div>

<div class="tabs">
  <div class="tab active" onclick="showPanel('cases')">個案稽核</div>
  <div class="tab" onclick="showPanel('appts')">預約稽核</div>
  <div class="tab" onclick="showPanel('missing')">未建檔個案</div>
</div>

<div id="panel-cases" class="panel active">
  <div class="filters">
    <div>狀態：</div>
    <button class="filter-btn active" onclick="filterCases('all',this)">全部</button>
    <button class="filter-btn error" onclick="filterCases('error',this)">🔴 錯誤</button>
    <button class="filter-btn warning" onclick="filterCases('warning',this)">🟡 警告</button>
    <button class="filter-btn ok" onclick="filterCases('ok',this)">🟢 正常</button>
    <div style="margin-left:8px">類別：</div>
    <button class="filter-btn" onclick="filterCaseCat('',this)">全部</button>
    <button class="filter-btn" onclick="filterCaseCat('A',this)">A</button>
    <button class="filter-btn" onclick="filterCaseCat('B',this)">B</button>
    <div style="margin-left:8px">心理師：</div>
    <select id="therapist-filter" onchange="filterCaseTherapist(this.value)">
      <option value="">全部</option>
    </select>
    <input type="text" id="case-search" placeholder="搜尋姓名/編號…" oninput="renderCases()" style="min-width:140px">
  </div>
  <div class="count-info" id="case-count"></div>
  <div class="table-wrap">
    <table>
      <thead><tr>
        <th>列</th><th>個案編號</th><th>姓名</th><th>心理師</th>
        <th>類別</th><th>進案日</th><th>性別</th><th>身分證</th>
        <th>經費來源</th><th>狀態</th>
      </tr></thead>
      <tbody id="case-tbody"></tbody>
    </table>
  </div>
  <div class="pagination" id="case-pages"></div>
</div>

<div id="panel-appts" class="panel">
  <div class="filters">
    <div>狀態：</div>
    <button class="filter-btn active" onclick="filterAppts('all',this)">全部</button>
    <button class="filter-btn error" onclick="filterAppts('error',this)">🔴 錯誤</button>
    <button class="filter-btn warning" onclick="filterAppts('warning',this)">🟡 警告</button>
    <button class="filter-btn ok" onclick="filterAppts('ok',this)">🟢 正常</button>
    <select id="appt-therapist" onchange="renderAppts()">
      <option value="">全部心理師</option>
    </select>
    <input type="text" id="appt-search" placeholder="搜尋姓名/編號…" oninput="renderAppts()" style="min-width:140px">
  </div>
  <div class="count-info" id="appt-count"></div>
  <div class="table-wrap">
    <table>
      <thead><tr>
        <th>列</th><th>日期</th><th>時間</th><th>場地</th>
        <th>心理師</th><th>個案</th><th>個案編號</th><th>備註</th><th>狀態</th>
      </tr></thead>
      <tbody id="appt-tbody"></tbody>
    </table>
  </div>
  <div class="pagination" id="appt-pages"></div>
  <div class="legend">⚪ 略過（個案編號空白）：<span id="skipped-count"></span> 筆</div>
</div>

<div id="panel-missing" class="panel">
  <p style="margin-bottom:16px; color:#555">以下個案有預約記錄但<strong>完全沒有個案基本資料</strong>，需聯絡對應心理師補建。</p>
  <h3 style="margin-bottom:8px; font-size:15px">❗ Category C — 預約 ≥ 3 次（優先處理）</h3>
  <div class="table-wrap" style="margin-bottom:24px">
    <table><thead><tr><th>代號</th><th>姓名</th><th>預約次數</th><th>心理師</th><th>建議行動</th></tr></thead>
    <tbody id="cat-c-tbody"></tbody></table>
  </div>
  <h3 style="margin-bottom:8px; font-size:15px">⚪ Category D — 預約 &lt; 3 次（低優先）</h3>
  <div class="table-wrap">
    <table><thead><tr><th>代號</th><th>姓名</th><th>預約次數</th><th>心理師</th><th>備註</th></tr></thead>
    <tbody id="cat-d-tbody"></tbody></table>
  </div>
</div>

<script>
const DATA = __DATA_PLACEHOLDER__;

let caseFilter = 'all';
let caseCatFilter = '';
let caseTherapistFilter = '';
let casePage = 1;
const CASE_PER_PAGE = 50;

let apptFilter = 'all';
let apptPage = 1;
const APPT_PER_PAGE = 100;

document.getElementById('gen-time').textContent = new Date().toLocaleString('zh-TW');

// Stats
const cases = DATA.cases;
const appts = DATA.appointments;
const sb = document.getElementById('stats-bar');
const cErr = cases.filter(r=>r.status==='error').length;
const cWarn = cases.filter(r=>r.status==='warning').length;
const cOk = cases.filter(r=>r.status==='ok').length;
const aErr = appts.filter(r=>r.status==='error').length;
const aWarn = appts.filter(r=>r.status==='warning').length;
const aOk = appts.filter(r=>r.status==='ok').length;

function card(label, value, cls) {
  return `<div class="stat-card ${cls}"><div class="label">${label}</div><div class="value">${value}</div></div>`;
}
sb.innerHTML =
  card('個案總計', cases.length, 'info') +
  card('個案 🔴 錯誤', cErr, 'error') +
  card('個案 🟡 警告', cWarn, 'warning') +
  card('個案 🟢 正常', cOk, 'ok') +
  card('預約總計', appts.length + DATA.skipped, 'info') +
  card('預約可匯入', aOk + aWarn, 'ok') +
  card('預約 🔴 錯誤', aErr, 'error') +
  card('預約略過', DATA.skipped, '');

// Tab
function showPanel(id) {
  document.querySelectorAll('.panel').forEach(p=>p.classList.remove('active'));
  document.querySelectorAll('.tab').forEach(t=>t.classList.remove('active'));
  document.getElementById('panel-'+id).classList.add('active');
  event.target.classList.add('active');
}

// Therapist dropdowns
const therapists = [...new Set(cases.map(r=>r.therapist).filter(Boolean))].sort();
const tSel = document.getElementById('therapist-filter');
therapists.forEach(t=>{ const o=document.createElement('option'); o.value=t; o.textContent=t; tSel.appendChild(o); });
const aSel = document.getElementById('appt-therapist');
const aTherapists = [...new Set(appts.map(r=>r.therapist).filter(Boolean))].sort();
aTherapists.forEach(t=>{ const o=document.createElement('option'); o.value=t; o.textContent=t; aSel.appendChild(o); });

// Filter cases
function filterCases(status, btn) {
  caseFilter = status; casePage = 1;
  document.querySelectorAll('#panel-cases .filter-btn').forEach(b=>b.classList.remove('active'));
  btn.classList.add('active');
  renderCases();
}
function filterCaseCat(cat, btn) {
  caseCatFilter = cat; casePage = 1;
  const btns = btn.parentElement.querySelectorAll('.filter-btn');
  btns.forEach(b=>{if(b.textContent.trim()===btn.textContent.trim())b.classList.add('active');else if(['A','B','全部'].some(x=>b.textContent.trim()===x))b.classList.remove('active');});
  renderCases();
}
function filterCaseTherapist(v) { caseTherapistFilter = v; casePage = 1; renderCases(); }

function getFilteredCases() {
  const q = document.getElementById('case-search').value.toLowerCase();
  return cases.filter(r => {
    if (caseFilter !== 'all' && r.status !== caseFilter) return false;
    if (caseCatFilter && r.category !== caseCatFilter) return false;
    if (caseTherapistFilter && r.therapist !== caseTherapistFilter) return false;
    if (q && !r.name.toLowerCase().includes(q) && !r.case_code.toLowerCase().includes(q)) return false;
    return true;
  });
}

function renderCases() {
  const filtered = getFilteredCases();
  const total = filtered.length;
  const pages = Math.ceil(total / CASE_PER_PAGE);
  if (casePage > pages) casePage = 1;
  const start = (casePage-1)*CASE_PER_PAGE;
  const slice = filtered.slice(start, start+CASE_PER_PAGE);

  document.getElementById('case-count').textContent = `顯示 ${start+1}–${Math.min(start+CASE_PER_PAGE,total)} / 共 ${total} 筆`;

  const tbody = document.getElementById('case-tbody');
  tbody.innerHTML = slice.map(r => {
    const issueHtml = r.issues.map(i=>`<div class="issue ${i.level}">${i.level==='error'?'🔴':'🟡'} [${i.field}] ${i.msg}</div>`).join('');
    const idShort = r.national_id ? r.national_id.substring(0,4)+'*****' : '—';
    return `<tr class="${r.status}">
      <td>${r.row}</td>
      <td><code>${r.case_code||'—'}</code></td>
      <td>${r.name||'—'}</td>
      <td>${r.therapist} <span style="color:#888;font-size:11px">${r.therapist_code||'?'}</span></td>
      <td><span class="badge cat-${r.category}">${r.category}</span></td>
      <td>${r.visit_date||'—'}</td>
      <td>${r.gender||'—'}</td>
      <td>${idShort}</td>
      <td title="${r.funding_raw}">${r.funding_source==='self_pay'?'自費':(r.institution||r.funding_raw||'—')}</td>
      <td><span class="badge ${r.status}">${r.status==='error'?'🔴錯誤':r.status==='warning'?'🟡警告':'🟢OK'}</span>
          <div class="issues">${issueHtml}</div></td>
    </tr>`;
  }).join('');

  renderPages('case-pages', pages, casePage, (p)=>{casePage=p; renderCases();});
}

// Filter appts
function filterAppts(status, btn) {
  apptFilter = status; apptPage = 1;
  document.querySelectorAll('#panel-appts .filter-btn').forEach(b=>b.classList.remove('active'));
  btn.classList.add('active');
  renderAppts();
}

function getFilteredAppts() {
  const q = document.getElementById('appt-search').value.toLowerCase();
  const t = document.getElementById('appt-therapist').value;
  return appts.filter(r => {
    if (apptFilter !== 'all' && r.status !== apptFilter) return false;
    if (t && r.therapist !== t) return false;
    if (q && !r.case_name.toLowerCase().includes(q) && !r.case_code.toLowerCase().includes(q) && !r.std_name.toLowerCase().includes(q)) return false;
    return true;
  });
}

function renderAppts() {
  const filtered = getFilteredAppts();
  const total = filtered.length;
  const pages = Math.ceil(total / APPT_PER_PAGE);
  if (apptPage > pages) apptPage = 1;
  const start = (apptPage-1)*APPT_PER_PAGE;
  const slice = filtered.slice(start, start+APPT_PER_PAGE);

  document.getElementById('appt-count').textContent = `顯示 ${start+1}–${Math.min(start+APPT_PER_PAGE,total)} / 共 ${total} 筆（略過 ${DATA.skipped} 筆）`;
  document.getElementById('skipped-count').textContent = DATA.skipped;

  const tbody = document.getElementById('appt-tbody');
  tbody.innerHTML = slice.map(r => {
    const issueHtml = r.issues.map(i=>`<div class="issue ${i.level}">${i.level==='error'?'🔴':'🟡'} [${i.field}] ${i.msg}</div>`).join('');
    return `<tr class="${r.status}">
      <td>${r.row}</td>
      <td>${r.date}</td>
      <td>${r.start}–${r.end}</td>
      <td>${r.room||'—'}</td>
      <td>${r.therapist} <span style="color:#888;font-size:11px">${r.therapist_code||'?'}</span></td>
      <td>${r.case_name}</td>
      <td><code>${r.case_code}</code></td>
      <td>${r.notes||''}</td>
      <td><span class="badge ${r.status}">${r.status==='error'?'🔴錯誤':r.status==='warning'?'🟡警告':'🟢OK'}</span>
          <div class="issues">${issueHtml}</div></td>
    </tr>`;
  }).join('');

  renderPages('appt-pages', pages, apptPage, (p)=>{apptPage=p; renderAppts();});
}

function renderPages(id, pages, current, onPage) {
  const el = document.getElementById(id);
  if (pages <= 1) { el.innerHTML=''; return; }
  let html = '';
  const range = 3;
  for (let p=1; p<=pages; p++) {
    if (p===1||p===pages||Math.abs(p-current)<=range) {
      html += `<button class="page-btn${p===current?' active':''}" onclick="(${onPage})(${p})">${p}</button>`;
    } else if (Math.abs(p-current)===range+1) {
      html += `<span style="padding:5px">…</span>`;
    }
  }
  el.innerHTML = html;
}

// Missing cases
const catC = DATA.cat_c;
const catD = DATA.cat_d;
document.getElementById('cat-c-tbody').innerHTML = catC.map(r=>
  `<tr><td><code>${r.code}</code></td><td>${r.name}</td><td>${r.appt_count}</td><td>${r.therapist}</td><td>${r.note}</td></tr>`
).join('');
document.getElementById('cat-d-tbody').innerHTML = catD.map(r=>
  `<tr><td><code>${r.code}</code></td><td>${r.name}</td><td>${r.appt_count}</td><td>${r.therapist}</td><td>${r.note}</td></tr>`
).join('');

// Init
renderCases();
renderAppts();
</script>
</body>
</html>
"""

# ─── 主程式 ───────────────────────────────────────────────────────────────────

def main() -> None:
    here = Path(__file__).parent.parent
    default_xlsx = here.parent.parent.parent / "2026年個案資料稽核報告_v2.xlsx"

    parser = argparse.ArgumentParser(description="慈恩資料遷移稽核工具")
    parser.add_argument("--xlsx", default=str(default_xlsx), help="xlsx 路徑")
    parser.add_argument("--output", default=str(here / "output" / "migration_report.html"), help="HTML 輸出路徑")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        print(f"找不到 xlsx 檔案：{xlsx_path}", file=sys.stderr)
        print(f"請用 --xlsx 指定路徑，例如：", file=sys.stderr)
        print(f"  python3.11 scripts/xlsx_audit.py --xlsx /path/to/2026年個案資料稽核報告_v2.xlsx", file=sys.stderr)
        sys.exit(1)

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    print(f"讀取：{xlsx_path}")
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)

    print("處理 Sheet3：進行中個案完整資料 …")
    cases = read_cases(wb["進行中個案完整資料"])

    print("處理 Sheet2：需處理個案清單 …")
    cat_c, cat_d, cat_b = read_problem_cases(wb["需處理個案清單"])

    print("處理 Sheet5：預約明細_標準化名稱 …")
    case_codes = {c["case_code"] for c in cases if c["case_code"]}
    # Add known codes from problem sheets
    case_codes |= {c["code"] for c in cat_c + cat_d + cat_b}
    appts, skipped = read_appointments(wb["預約明細_標準化名稱"], case_codes)

    # Summary
    c_err = sum(1 for c in cases if c["status"] == "error")
    c_warn = sum(1 for c in cases if c["status"] == "warning")
    c_ok = sum(1 for c in cases if c["status"] == "ok")
    a_err = sum(1 for a in appts if a["status"] == "error")
    a_warn = sum(1 for a in appts if a["status"] == "warning")
    a_ok = sum(1 for a in appts if a["status"] == "ok")

    print(f"\n{'─'*50}")
    print(f"個案：{len(cases)} 筆　🔴{c_err} 錯誤　🟡{c_warn} 警告　🟢{c_ok} 正常")
    print(f"預約：{len(appts)} 筆（+略過 {skipped}）　🔴{a_err} 錯誤　🟡{a_warn} 警告　🟢{a_ok} 正常")
    print(f"未建檔：Category C {len(cat_c)} 人　Category D {len(cat_d)} 人")

    data = {
        "cases": cases,
        "appointments": appts,
        "skipped": skipped,
        "cat_c": cat_c,
        "cat_d": cat_d,
        "cat_b": cat_b,
    }

    html = HTML_TEMPLATE.replace("__DATA_PLACEHOLDER__", json.dumps(data, ensure_ascii=False, default=str))
    output_path.write_text(html, encoding="utf-8")
    print(f"\n稽核報告已輸出：{output_path}")
    print("請用瀏覽器開啟該檔案。")


if __name__ == "__main__":
    main()
