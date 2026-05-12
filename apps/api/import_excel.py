"""Import data from Excel files into the database.

Usage:
    cd apps/api
    PYTHONPATH=. python import_excel.py --cases /path/to/慈恩個案管理總表.xlsx
    PYTHONPATH=. python import_excel.py --ledger /path/to/慈恩會計行政月報表日更格式1150503.xlsx
    PYTHONPATH=. python import_excel.py --all /path/to/cases.xlsx /path/to/ledger.xlsx
"""

import argparse
import re
from datetime import date, datetime
from decimal import Decimal

import openpyxl

from app.database import SessionLocal
from app.models.case import Case
from app.models.institution import Institution
from app.models.room import Room
from app.models.session_record import SessionRecord
from app.models.user import User


def get_therapist_map(db) -> dict[str, int]:
    users = db.query(User).filter(User.role == "therapist").all()
    return {u.name: u.id for u in users}


def get_room_map(db) -> dict[str, int]:
    rooms = db.query(Room).all()
    return {r.room_code: r.id for r in rooms}


def truncate(val, maxlen: int) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    if not s or s == "None":
        return None
    return s[:maxlen]


def parse_roc_date(val) -> date | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    if not s:
        return None
    m = re.match(r"(\d+)/(\d+)/(\d+)", s)
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if y < 200:
            y += 1911
        try:
            return date(y, mo, d)
        except ValueError:
            return None
    m = re.match(r"(\d+)\.(\d+)\.(\d+)", s)
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if y < 200:
            y += 1911
        try:
            return date(y, mo, d)
        except ValueError:
            return None
    return None


def parse_gender(val) -> str | None:
    if val is None:
        return None
    v = str(val).strip()
    if v in ("1", "1.0", "男"):
        return "male"
    if v in ("2", "2.0", "女"):
        return "female"
    return None


def resolve_therapist(name: str | None, therapist_map: dict[str, int]) -> int | None:
    if not name:
        return None
    name = name.strip()
    tid = therapist_map.get(name)
    if tid:
        return tid
    for tname, t_id in therapist_map.items():
        if tname in name or name in tname:
            return t_id
    return None


def detect_column_map(ws) -> tuple[dict[str, int], int]:
    """Auto-detect header row and column positions.

    Returns (column_map, data_start_row).
    Newer sheets (109+): header at row 1, data from row 2.
    Older sheets (107, 108): header at row 4, data from row 5.
    """
    header_candidates = [1, 4]
    for hr in header_candidates:
        headers = [str(c.value).strip() if c.value else "" for c in ws[hr]]
        if "個案姓名" in headers:
            col_map = {}
            for i, h in enumerate(headers):
                if h:
                    col_map[h] = i
            return col_map, hr + 1
    return {}, 2


# ---------------------------------------------------------------------------
# Cases import
# ---------------------------------------------------------------------------

def get_or_create_institution(db, name: str, inst_cache: dict[str, int]) -> int:
    if name in inst_cache:
        return inst_cache[name]
    existing = db.query(Institution).filter(Institution.name == name).first()
    if existing:
        if not existing.is_active:
            existing.is_active = True
            db.flush()
        inst_cache[name] = existing.id
        return existing.id
    inst = Institution(name=name)
    db.add(inst)
    db.flush()
    inst_cache[name] = inst.id
    return inst.id


def import_cases(filepath: str):
    db = SessionLocal()
    therapist_map = get_therapist_map(db)
    inst_cache: dict[str, int] = {}

    existing = db.query(Case).filter(Case.case_code.isnot(None)).count()
    if existing > 0:
        print(f"Clearing {existing} previously imported cases...")
        db.query(Case).filter(Case.case_code.isnot(None)).delete()
        db.commit()

    wb = openpyxl.load_workbook(filepath, data_only=True)
    total = 0

    allowed_sheets = ["115年個案資料", "114年個案資料"]

    for sheet_name in wb.sheetnames:
        if sheet_name not in allowed_sheets:
            print(f"  {sheet_name}: skipped (before 114)")
            continue

        ws = wb[sheet_name]
        col_map, data_start = detect_column_map(ws)

        if not col_map:
            print(f"  {sheet_name}: skipped (no header found)")
            continue

        def col(name: str, row_vals) -> str | None:
            idx = col_map.get(name)
            if idx is None or idx >= len(row_vals):
                return None
            return row_vals[idx]

        count = 0
        for row in ws.iter_rows(min_row=data_start, max_row=ws.max_row, values_only=True):
            name = truncate(col("個案姓名", row), 100)
            if not name:
                continue

            therapist_id = resolve_therapist(
                truncate(col("接案心理師", row), 100), therapist_map
            )
            if not therapist_id:
                therapist_id = 1

            funding_raw = truncate(col("經費來源", row), 100) or "自費"
            if funding_raw == "自費":
                funding_source = "self_pay"
                institution_id = None
            else:
                funding_source = "institution"
                institution_id = get_or_create_institution(db, funding_raw, inst_cache)

            raw_date = col("進案日", row)
            if raw_date is None:
                raw_date = col("進案年月", row)
            initial_visit = None
            if isinstance(raw_date, (datetime, date)):
                initial_visit = raw_date if isinstance(raw_date, date) else raw_date.date()
            elif raw_date:
                initial_visit = parse_roc_date(str(raw_date))

            birth_raw = col("出生日期", row)
            birth_date = parse_roc_date(birth_raw)

            gender_raw = col("性別", row)
            case_code_raw = col("個案編號", row)
            if case_code_raw is None:
                idx0 = col_map.get("個案編號")
                if idx0 is None:
                    for k in col_map:
                        if k == "" or k is None:
                            break
                    case_code_raw = row[0] if row[0] else None

            case = Case(
                case_code=truncate(case_code_raw, 20) if case_code_raw else None,
                name=name,
                birth_date=birth_date,
                gender=parse_gender(gender_raw),
                phone=truncate(col("連絡電話(手機)", row), 50),
                phone_home=truncate(col("連絡電話(市話)", row), 50),
                address=truncate(col("地址", row), 500),
                emergency_contact=truncate(col("緊急聯絡人", row), 200),
                emergency_phone=truncate(col("聯絡電話1", row), 50),
                emergency_phone2=truncate(col("聯絡電話2", row), 50),
                initial_visit_date=initial_visit,
                funding_source=funding_source,
                institution_id=institution_id,
                referral_source=truncate(col("轉介來源", row), 200),
                session_location=truncate(col("會談地點", row), 200),
                therapist_id=therapist_id,
                status="ongoing",
                notes=f"[{sheet_name}]",
            )
            db.add(case)
            count += 1

        print(f"  {sheet_name}: {count} cases")
        total += count

    db.commit()
    print(f"Institutions created: {list(inst_cache.keys())}")
    db.close()
    print(f"Total: imported {total} cases")


# ---------------------------------------------------------------------------
# Ledger (raw sheet) import
# ---------------------------------------------------------------------------

def import_ledger(filepath: str):
    db = SessionLocal()
    therapist_map = get_therapist_map(db)
    room_map = get_room_map(db)

    existing = db.query(SessionRecord).filter(SessionRecord.appointment_id.is_(None)).count()
    if existing > 0:
        print(f"Clearing {existing} previously imported ledger records...")
        db.query(SessionRecord).filter(SessionRecord.appointment_id.is_(None)).delete()
        db.commit()

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["raw"]

    imported = 0
    skipped = 0

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        dt = row[1]
        if not isinstance(dt, (date, datetime)):
            skipped += 1
            continue

        session_date = dt.date() if isinstance(dt, datetime) else dt

        therapist_id = resolve_therapist(str(row[3]) if row[3] else None, therapist_map)
        if not therapist_id:
            skipped += 1
            continue

        room_code = truncate(row[2], 10)
        room_id = room_map.get(room_code) if room_code else None

        amount_val = row[16] or row[15] or row[5]
        try:
            amount = Decimal(str(amount_val)) if amount_val else Decimal("0")
        except Exception:
            amount = Decimal("0")

        status_raw = truncate(row[8], 20) or "未收"
        funding = truncate(row[4], 100)
        if status_raw == "已收":
            payment_status = "paid"
        elif funding and funding != "自費":
            payment_status = "pending_claim"
        else:
            payment_status = "unpaid"

        session_format = truncate(row[12], 20)
        location = truncate(row[7], 20)
        if session_format and "視訊" in session_format:
            session_type = "online"
        elif location and "到宅" in location:
            session_type = "home_visit"
        else:
            session_type = "in_person"

        record = SessionRecord(
            appointment_id=None,
            session_date=session_date,
            case_id=None,
            therapist_id=therapist_id,
            session_type=session_type,
            room_id=room_id,
            fee_category="counseling",
            amount=amount,
            payment_status=payment_status,
        )
        db.add(record)
        imported += 1

    db.commit()
    db.close()
    print(f"Ledger: imported {imported} records, skipped {skipped}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Import Excel data into CheerPsy")
    parser.add_argument("--cases", help="Path to 個案管理總表.xlsx")
    parser.add_argument("--ledger", help="Path to 會計行政月報表.xlsx")
    parser.add_argument("--all", nargs=2, metavar=("CASES", "LEDGER"), help="Import both files")
    args = parser.parse_args()

    if args.all:
        print("=== Importing cases ===")
        import_cases(args.all[0])
        print("\n=== Importing ledger ===")
        import_ledger(args.all[1])
    elif args.cases:
        import_cases(args.cases)
    elif args.ledger:
        import_ledger(args.ledger)
    else:
        parser.print_help()
