import io
from datetime import date
from decimal import Decimal

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import extract, func
from sqlalchemy.orm import Session, joinedload

from app.auth.dependencies import RequireRole, get_current_user
from app.database import get_db
from app.models.appointment import Appointment
from app.models.case import Case
from app.models.petty_cash import PettyCash
from app.models.product_sales import ProductSale
from app.models.session_record import SessionRecord
from app.models.user import User

router = APIRouter(prefix="/reports", tags=["reports"])

DEFAULT_COMMISSION_RATE = Decimal("0.70")


def _get_rate(r: SessionRecord) -> Decimal:
    if r.commission_rate_used is not None:
        return Decimal(str(r.commission_rate_used))
    return DEFAULT_COMMISSION_RATE


def _eff(r: SessionRecord) -> float:
    """Effective amount = original amount minus 優待 discount."""
    return float(r.amount) - float(r.discount_amount or 0)


@router.get("/monthly")
def monthly_report(
    year: int = Query(...),
    month: int = Query(...),
    user: User = Depends(RequireRole(["admin", "accountant"])),
    db: Session = Depends(get_db),
):
    records = (
        db.query(SessionRecord)
        .filter(
            extract("year", SessionRecord.session_date) == year,
            extract("month", SessionRecord.session_date) == month,
            SessionRecord.is_void.is_(False),
        )
        .all()
    )

    total_revenue = sum(_eff(r) for r in records)
    total_therapist = sum(round(_eff(r) * float(_get_rate(r)), 2) for r in records)
    total_clinic = sum(round(_eff(r) * float(1 - _get_rate(r)), 2) for r in records)
    session_count = len(records)

    paid_count = sum(1 for r in records if r.payment_status in ("paid", "claimed"))
    unpaid_count = sum(1 for r in records if r.payment_status in ("unpaid", "claiming"))

    # Session type breakdown
    by_type: dict[str, dict] = {}
    for r in records:
        t = r.session_type or "unknown"
        if t not in by_type:
            by_type[t] = {"count": 0, "revenue": 0.0}
        by_type[t]["count"] += 1
        by_type[t]["revenue"] += _eff(r)

    # Funding source breakdown (self_pay vs institution advance)
    self_pay_revenue = 0.0
    institution_revenue = 0.0
    institution_unpaid = 0.0
    for r in records:
        case = None
        if r.appointment:
            case = r.appointment.case
        elif r.case_id:
            case = db.query(Case).filter(Case.id == r.case_id).first()

        funding = r.funding_source or (case.funding_source if case else "self_pay")
        if funding == "institution":
            institution_revenue += _eff(r)
            if r.payment_status in ("unpaid", "claiming"):
                institution_unpaid += _eff(r)
        else:
            self_pay_revenue += _eff(r)

    # Therapist summary
    by_therapist: dict[int, dict] = {}
    for r in records:
        tid = r.therapist_id
        if tid not in by_therapist:
            by_therapist[tid] = {"therapist_id": tid, "sessions": 0, "revenue": 0.0, "share": 0.0}
        by_therapist[tid]["sessions"] += 1
        by_therapist[tid]["revenue"] += _eff(r)
        by_therapist[tid]["share"] += round(_eff(r) * float(_get_rate(r)), 2)

    therapist_names = {}
    if by_therapist:
        users = db.query(User).filter(User.id.in_(by_therapist.keys())).all()
        therapist_names = {u.id: u.name for u in users}

    therapist_summary = []
    for tid, data in by_therapist.items():
        therapist_summary.append({
            "therapist_id": tid,
            "therapist_name": therapist_names.get(tid, ""),
            "sessions": data["sessions"],
            "revenue": data["revenue"],
            "therapist_share": data["share"],
            "clinic_share": round(data["revenue"] - data["share"], 2),
        })
    therapist_summary.sort(key=lambda x: x["revenue"], reverse=True)

    # Petty cash
    petty = (
        db.query(PettyCash)
        .filter(
            extract("year", PettyCash.date) == year,
            extract("month", PettyCash.date) == month,
        )
        .all()
    )
    petty_total = sum(float(p.amount) for p in petty)
    by_category: dict[str, float] = {}
    for p in petty:
        by_category[p.category] = by_category.get(p.category, 0) + float(p.amount)

    # Other product sales (separate from session revenue)
    product_rows = (
        db.query(ProductSale)
        .filter(
            extract("year", ProductSale.sale_date) == year,
            extract("month", ProductSale.sale_date) == month,
            ProductSale.is_void.is_(False),
        )
        .all()
    )
    product_total = sum(float(p.amount) * p.quantity for p in product_rows)
    product_by_method: dict[str, float] = {}
    for p in product_rows:
        m = p.payment_method or "cash"
        product_by_method[m] = product_by_method.get(m, 0) + float(p.amount) * p.quantity

    # Appointment stats
    month_appts = (
        db.query(Appointment)
        .filter(
            extract("year", Appointment.created_at) == year,
            extract("month", Appointment.created_at) == month,
        )
        .all()
    )
    booked = sum(1 for a in month_appts if a.status == "booked")
    cancelled = sum(1 for a in month_appts if a.status == "cancelled")
    cancelled_amount = sum(float(a.amount) for a in month_appts if a.status == "cancelled")

    # KPIs
    total_appts = booked + session_count + cancelled
    cancel_rate = round(cancelled / total_appts * 100, 1) if total_appts > 0 else 0

    # Continuation rate: cases with >1 session this month
    case_session_counts: dict[int, int] = {}
    for r in records:
        if r.case_id:
            case_session_counts[r.case_id] = case_session_counts.get(r.case_id, 0) + 1
    cases_with_sessions = len(case_session_counts)
    continuing_cases = sum(1 for c in case_session_counts.values() if c > 1)
    continuation_rate = round(continuing_cases / cases_with_sessions * 100, 1) if cases_with_sessions > 0 else 0

    return {
        "year": year,
        "month": month,
        "summary": {
            "total_revenue": total_revenue,
            "total_therapist_share": total_therapist,
            "total_clinic_share": total_clinic,
            "session_count": session_count,
            "paid_count": paid_count,
            "unpaid_count": unpaid_count,
            "petty_cash_total": petty_total,
            "net_clinic_income": total_clinic + petty_total,
        },
        "product_sales": {
            "total": product_total,
            "count": len(product_rows),
            "by_payment_method": product_by_method,
        },
        "session_type_breakdown": by_type,
        "funding_breakdown": {
            "self_pay_revenue": self_pay_revenue,
            "institution_revenue": institution_revenue,
            "institution_unpaid": institution_unpaid,
        },
        "pnl": {
            "gross_revenue": total_revenue,
            "therapist_cost": total_therapist,
            "clinic_gross": total_clinic,
            "petty_cash_expense": petty_total,
            "net_income": total_clinic + petty_total,
            "cancelled_loss": cancelled_amount,
        },
        "kpi": {
            "cancel_rate": cancel_rate,
            "continuation_rate": continuation_rate,
            "active_therapists": len(by_therapist),
            "unique_cases": cases_with_sessions,
        },
        "therapist_summary": therapist_summary,
        "petty_cash_by_category": by_category,
        "appointment_stats": {
            "booked": booked,
            "cancelled": cancelled,
            "executed": session_count,
        },
    }


@router.get("/reconciliation")
def reconciliation_report(
    year: int = Query(...),
    month: int = Query(...),
    user: User = Depends(RequireRole(["admin", "accountant"])),
    db: Session = Depends(get_db),
):
    records = (
        db.query(SessionRecord)
        .filter(
            extract("year", SessionRecord.session_date) == year,
            extract("month", SessionRecord.session_date) == month,
            SessionRecord.is_void.is_(False),
        )
        .order_by(SessionRecord.session_date)
        .all()
    )

    self_pay_records = []
    institution_groups: dict[str, list] = {}

    for r in records:
        case = None
        if r.appointment:
            case = r.appointment.case
        elif r.case_id:
            case = db.query(Case).filter(Case.id == r.case_id).first()

        funding = r.funding_source or (case.funding_source if case else "self_pay")
        institution_name = None
        if case and case.institution:
            institution_name = case.institution.name

        therapist = None
        if r.appointment:
            therapist = r.appointment.therapist
        else:
            therapist = db.query(User).filter(User.id == r.therapist_id).first()

        row = {
            "id": r.id,
            "session_date": r.session_date.isoformat(),
            "case_id": r.case_id,
            "case_name": case.name if case else None,
            "therapist_name": therapist.name if therapist else None,
            "amount": _eff(r),
            "payment_status": r.payment_status,
            "claim_number": r.claim_number,
            "receipt_number": r.receipt_number,
        }

        if funding == "institution":
            key = institution_name or "未知機構"
            institution_groups.setdefault(key, []).append(row)
        else:
            self_pay_records.append(row)

    # Summaries
    def summarize(recs):
        total = sum(r["amount"] for r in recs)
        paid = sum(r["amount"] for r in recs if r["payment_status"] in ("paid", "claimed"))
        unpaid = sum(r["amount"] for r in recs if r["payment_status"] in ("unpaid", "claiming"))
        return {"total": total, "paid": paid, "unpaid": unpaid, "count": len(recs)}

    institution_summaries = []
    for name, recs in sorted(institution_groups.items()):
        institution_summaries.append({
            "institution_name": name,
            "summary": summarize(recs),
            "records": recs,
        })

    return {
        "year": year,
        "month": month,
        "self_pay": {
            "summary": summarize(self_pay_records),
            "records": self_pay_records,
        },
        "institutions": institution_summaries,
    }


# ---------------------------------------------------------------------------
# 進案統計 helpers
# ---------------------------------------------------------------------------

def _age_group(birth_date, case_age, ref_date: date) -> str:
    """Return 'child' if age < 18 at ref_date, else 'adult'."""
    if birth_date is not None:
        age = (ref_date - birth_date).days // 365
    elif case_age is not None:
        age = case_age
    else:
        return "adult"
    return "child" if age < 18 else "adult"


def _intake_data(db: Session, year: int, month: int) -> dict:
    """
    Return structured intake statistics for the given year/month.
    A "new intake" = activated case (has case_number) whose first non-voided
    session falls in the specified month.
    """
    # Subquery: earliest session date per case (exclude voided)
    first_sq = (
        db.query(
            SessionRecord.case_id.label("case_id"),
            func.min(SessionRecord.session_date).label("first_date"),
        )
        .filter(SessionRecord.is_void.is_(False))
        .group_by(SessionRecord.case_id)
        .subquery()
    )

    # Cases whose first session is in the target month and are activated
    rows = (
        db.query(Case, first_sq.c.first_date)
        .join(first_sq, Case.id == first_sq.c.case_id)
        .filter(
            extract("year", first_sq.c.first_date) == year,
            extract("month", first_sq.c.first_date) == month,
            Case.case_number.isnot(None),
        )
        .all()
    )

    # Eagerly load therapist / institution to avoid N+1
    case_ids = [c.id for c, _ in rows]
    if case_ids:
        cases_with_relations = (
            db.query(Case)
            .options(joinedload(Case.therapist), joinedload(Case.institution))
            .filter(Case.id.in_(case_ids))
            .all()
        )
        case_map = {c.id: c for c in cases_with_relations}
    else:
        case_map = {}

    therapist_map: dict[int, dict] = {}
    institution_map: dict[str, dict] = {}
    summary = {
        "institution_adult": 0,
        "institution_child": 0,
        "self_pay_adult": 0,
        "self_pay_child": 0,
        "total": 0,
    }
    case_list = []

    for c_stub, first_date in rows:
        c = case_map.get(c_stub.id, c_stub)
        ag = _age_group(c.birth_date, c.age, first_date)
        funding = c.funding_source or "self_pay"
        is_inst = funding == "institution"

        therapist_name = c.therapist.name if c.therapist else "（未指定）"
        therapist_id = c.therapist_id or 0
        institution_name = c.institution.name if c.institution else None

        # Summary
        key = ("institution" if is_inst else "self_pay") + "_" + ag
        summary[key] += 1
        summary["total"] += 1

        # By therapist
        if therapist_id not in therapist_map:
            therapist_map[therapist_id] = {
                "therapist_name": therapist_name,
                "institution_adult": 0,
                "institution_child": 0,
                "self_pay_adult": 0,
                "self_pay_child": 0,
                "designated_institution_adult": 0,
                "designated_institution_child": 0,
                "designated_self_pay_adult": 0,
                "designated_self_pay_child": 0,
                "total": 0,
            }
        td = therapist_map[therapist_id]
        td[key] += 1
        td["total"] += 1
        if c.is_designated:
            td["designated_" + key] += 1

        # By institution
        if is_inst and institution_name:
            if institution_name not in institution_map:
                institution_map[institution_name] = {"institution_name": institution_name, "adult": 0, "child": 0, "total": 0}
            institution_map[institution_name][ag] += 1
            institution_map[institution_name]["total"] += 1

        case_list.append({
            "case_number": c.case_number,
            "name": c.name,
            "funding_source": funding,
            "age_group": ag,
            "therapist_name": therapist_name,
            "institution_name": institution_name,
            "is_designated": c.is_designated,
            "first_session_date": first_date.isoformat() if first_date else None,
        })

    by_therapist = sorted(therapist_map.values(), key=lambda x: x["total"], reverse=True)
    by_institution = sorted(institution_map.values(), key=lambda x: x["total"], reverse=True)

    return {
        "year": year,
        "month": month,
        "summary": summary,
        "by_therapist": by_therapist,
        "by_institution": by_institution,
        "cases": case_list,
    }


def _build_intake_excel(data: dict) -> io.BytesIO:
    wb = Workbook()
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDEEFF")
    center = Alignment(horizontal="center")

    # ── Sheet 1: 摘要 ──────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "摘要"
    year, month = data["year"], data["month"]
    ws1["A1"] = f"{year}年{month}月 新進案統計摘要"
    ws1["A1"].font = Font(bold=True, size=13)
    ws1.merge_cells("A1:C1")

    ws1.append([])
    ws1.append(["類別", "成人", "兒青（<18歲）", "小計"])
    for cell in ws1[ws1.max_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    s = data["summary"]
    ws1.append(["機構案", s["institution_adult"], s["institution_child"],
                s["institution_adult"] + s["institution_child"]])
    ws1.append(["自費案", s["self_pay_adult"], s["self_pay_child"],
                s["self_pay_adult"] + s["self_pay_child"]])
    ws1.append(["合計",
                s["institution_adult"] + s["self_pay_adult"],
                s["institution_child"] + s["self_pay_child"],
                s["total"]])
    for cell in ws1[ws1.max_row]:
        cell.font = Font(bold=True)

    for col in ["A", "B", "C", "D"]:
        ws1.column_dimensions[col].width = 18

    # ── Sheet 2: 各心理師 ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("各心理師")
    headers2 = [
        "心理師",
        "機構成人", "機構兒青",
        "自費成人", "自費兒青",
        "指定-機構成人", "指定-機構兒青",
        "指定-自費成人", "指定-自費兒青",
        "合計",
    ]
    ws2.append(headers2)
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for row in data["by_therapist"]:
        ws2.append([
            row["therapist_name"],
            row["institution_adult"], row["institution_child"],
            row["self_pay_adult"], row["self_pay_child"],
            row["designated_institution_adult"], row["designated_institution_child"],
            row["designated_self_pay_adult"], row["designated_self_pay_child"],
            row["total"],
        ])

    ws2.column_dimensions["A"].width = 20
    for col in ["B", "C", "D", "E", "F", "G", "H", "I", "J"]:
        ws2.column_dimensions[col].width = 14

    # ── Sheet 3: 各機構 ────────────────────────────────────────────────────
    ws3 = wb.create_sheet("各機構")
    headers3 = ["機構名稱", "成人", "兒青（<18歲）", "合計"]
    ws3.append(headers3)
    for cell in ws3[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for row in data["by_institution"]:
        ws3.append([row["institution_name"], row["adult"], row["child"], row["total"]])

    ws3.column_dimensions["A"].width = 24
    for col in ["B", "C", "D"]:
        ws3.column_dimensions[col].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# 進案統計 endpoints
# ---------------------------------------------------------------------------

@router.get("/intake")
def get_intake_stats(
    year: int = Query(...),
    month: int = Query(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    return _intake_data(db, year, month)


@router.get("/intake/export")
def export_intake_excel(
    year: int = Query(...),
    month: int = Query(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    data = _intake_data(db, year, month)
    buf = _build_intake_excel(data)
    filename = f"intake_{year}{month:02d}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
